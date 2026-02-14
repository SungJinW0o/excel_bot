import os

import pandas as pd
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .auth import authorize, get_user
from .config import load_config
from .events import emit_event
from .notifications import (
    notify_data_cleaned,
    notify_pipeline_completed,
    notify_pipeline_failed,
    notify_pipeline_started,
)


def _safe_divide(numerator: float, denominator: float) -> float:
    if not denominator:
        return 0.0
    return float(numerator / denominator)


def _write_data_quality_issues(output_dir: str, issues: list[dict]) -> str:
    issues_path = os.path.join(output_dir, "data_quality_issues.csv")
    pd.DataFrame(issues, columns=["File", "Issue"]).to_csv(issues_path, index=False)
    return issues_path


def _enrich_financial_columns(
    df: pd.DataFrame,
    qty_col: str,
    price_col: str,
    expense_col: str,
) -> pd.DataFrame:
    working = df.copy()
    working[qty_col] = pd.to_numeric(working[qty_col], errors="coerce")
    working[price_col] = pd.to_numeric(working[price_col], errors="coerce")

    if expense_col in working.columns:
        expense_series = pd.to_numeric(working[expense_col], errors="coerce").fillna(0.0)
    else:
        expense_series = pd.Series(0.0, index=working.index, dtype="float64")

    working["TotalRevenue"] = working[qty_col] * working[price_col]
    working["TotalExpense"] = expense_series
    working["Savings"] = working["TotalRevenue"] - working["TotalExpense"]
    return working


def _add_savings_rate(df: pd.DataFrame, earning_col: str, savings_col: str) -> pd.DataFrame:
    result = df.copy()
    result["SavingsRate"] = 0.0
    non_zero = result[earning_col] != 0
    result.loc[non_zero, "SavingsRate"] = (
        result.loc[non_zero, savings_col] / result.loc[non_zero, earning_col]
    )
    return result


def _deduplicate_cleaned_data(df: pd.DataFrame, order_id_col: str | None) -> pd.DataFrame:
    if df.empty:
        return df
    if order_id_col and order_id_col in df.columns:
        return df.drop_duplicates(subset=[order_id_col], keep="last")
    return df.drop_duplicates(keep="last")


def _auto_fit_columns(ws) -> None:
    for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        values = [str(c.value) for c in col_cells if c.value is not None]
        longest = max((len(v) for v in values), default=10)
        width = min(max(longest + 2, 12), 36)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = width


def _style_table_headers(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.freeze_panes = "A2"


def _apply_number_formats(ws, formats: dict) -> None:
    header_to_col = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value is not None:
            header_to_col[str(value)] = col

    for header, number_format in formats.items():
        col = header_to_col.get(header)
        if not col:
            continue
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = number_format


def _format_report_workbook(workbook) -> None:
    sheet_formats = {
        "Overall_Summary": {
            "TotalRevenue": "$#,##0.00",
            "TotalEarning": "$#,##0.00",
            "Expenses": "$#,##0.00",
            "Savings": "$#,##0.00",
            "SavingsRate": "0.00%",
            "AverageOrderValue": "$#,##0.00",
            "TotalOrders": "#,##0",
        },
        "Category_Summary": {
            "TotalRevenue": "$#,##0.00",
            "TotalEarning": "$#,##0.00",
            "Expenses": "$#,##0.00",
            "Savings": "$#,##0.00",
            "SavingsRate": "0.00%",
            "TotalQuantity": "#,##0",
        },
        "Region_Summary": {
            "TotalRevenue": "$#,##0.00",
            "TotalEarning": "$#,##0.00",
            "Expenses": "$#,##0.00",
            "Savings": "$#,##0.00",
            "SavingsRate": "0.00%",
            "TotalOrders": "#,##0",
        },
        "Benchmark_Overall": {
            "TotalEarning": "$#,##0.00",
            "Expenses": "$#,##0.00",
            "Savings": "$#,##0.00",
            "SavingsRate": "0.00%",
            "AverageOrderValue": "$#,##0.00",
            "TotalOrders": "#,##0",
        },
        "Benchmark_Category": {
            "TotalEarning": "$#,##0.00",
            "Expenses": "$#,##0.00",
            "Savings": "$#,##0.00",
            "SavingsRate": "0.00%",
            "TotalQuantity": "#,##0",
        },
        "Benchmark_Region": {
            "TotalEarning": "$#,##0.00",
            "Expenses": "$#,##0.00",
            "Savings": "$#,##0.00",
            "SavingsRate": "0.00%",
            "TotalOrders": "#,##0",
        },
    }

    for sheet_name, formats in sheet_formats.items():
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        _style_table_headers(ws)
        _apply_number_formats(ws, formats)
        _auto_fit_columns(ws)

    if "Benchmark_Metrics" in workbook.sheetnames:
        ws = workbook["Benchmark_Metrics"]
        _style_table_headers(ws)
        for row in range(2, ws.max_row + 1):
            metric_name = str(ws.cell(row=row, column=1).value or "")
            value_cell = ws.cell(row=row, column=2)
            if metric_name == "SavingsRate":
                value_cell.number_format = "0.00%"
            elif metric_name == "TotalOrders":
                value_cell.number_format = "#,##0"
            else:
                value_cell.number_format = "$#,##0.00"
        _auto_fit_columns(ws)


def _paint_kpi_card(ws, start_col: int, title: str, formula: str, fill_color: str, number_format: str) -> None:
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=start_col + 1)
    ws.merge_cells(start_row=5, start_column=start_col, end_row=6, end_column=start_col + 1)

    for row in range(4, 7):
        for col in range(start_col, start_col + 2):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    title_cell = ws.cell(row=4, column=start_col)
    title_cell.value = title
    title_cell.font = Font(color="FFFFFF", bold=True, size=11)

    value_cell = ws.cell(row=5, column=start_col)
    value_cell.value = formula
    value_cell.number_format = number_format
    value_cell.font = Font(color="FFFFFF", bold=True, size=14)


def _build_executive_dashboard(workbook, category_col: str, region_col: str) -> None:
    if "Executive_Dashboard" in workbook.sheetnames:
        del workbook["Executive_Dashboard"]

    ws = workbook.create_sheet("Executive_Dashboard")
    ws.sheet_view.showGridLines = False
    ws["B1"] = "Executive Financial Dashboard"
    ws["B1"].font = Font(size=24, bold=True, color="0F172A")
    ws["B2"] = "Auto-generated benchmark visuals for company-style reporting."
    ws["B2"].font = Font(size=11, color="475569")

    for col in range(2, 21):
        ws.column_dimensions[get_column_letter(col)].width = 12

    _paint_kpi_card(ws, 2, "Total Earning", "='Benchmark_Overall'!B2", "14532D", "$#,##0.00")
    _paint_kpi_card(ws, 5, "Expenses", "='Benchmark_Overall'!C2", "991B1B", "$#,##0.00")
    _paint_kpi_card(ws, 8, "Savings", "='Benchmark_Overall'!D2", "1D4ED8", "$#,##0.00")
    _paint_kpi_card(ws, 11, "Savings Rate", "='Benchmark_Overall'!E2", "9A3412", "0.00%")
    _paint_kpi_card(ws, 14, "Avg Order Value", "='Benchmark_Overall'!F2", "581C87", "$#,##0.00")
    _paint_kpi_card(ws, 17, "Total Orders", "='Benchmark_Overall'!A2", "0F766E", "#,##0")

    category_ws = workbook["Benchmark_Category"]
    category_max_row = max(category_ws.max_row, 2)
    category_bar = BarChart()
    category_bar.title = "Category Performance: Earnings vs Expenses"
    category_bar.y_axis.title = "Amount"
    category_bar.x_axis.title = category_col
    category_bar.style = 10
    category_data = Reference(category_ws, min_col=2, max_col=3, min_row=1, max_row=category_max_row)
    category_cats = Reference(category_ws, min_col=1, min_row=2, max_row=category_max_row)
    category_bar.add_data(category_data, titles_from_data=True)
    category_bar.set_categories(category_cats)
    category_bar.height = 8
    category_bar.width = 13
    ws.add_chart(category_bar, "B9")

    region_ws = workbook["Benchmark_Region"]
    region_max_row = max(region_ws.max_row, 2)
    region_pie = PieChart()
    region_pie.title = "Savings Share by Region"
    region_data = Reference(region_ws, min_col=5, min_row=1, max_row=region_max_row)
    region_cats = Reference(region_ws, min_col=1, min_row=2, max_row=region_max_row)
    region_pie.add_data(region_data, titles_from_data=True)
    region_pie.set_categories(region_cats)
    region_pie.dataLabels = DataLabelList()
    region_pie.dataLabels.showPercent = True
    region_pie.height = 8
    region_pie.width = 8.5
    ws.add_chart(region_pie, "N9")

    savings_rate_line = LineChart()
    savings_rate_line.title = "Savings Rate Trend by Category"
    savings_rate_line.y_axis.title = "Savings Rate"
    savings_rate_line.y_axis.number_format = "0%"
    savings_rate_line.x_axis.title = category_col
    savings_rate_line.style = 2
    line_data = Reference(category_ws, min_col=5, max_col=5, min_row=1, max_row=category_max_row)
    line_cats = Reference(category_ws, min_col=1, min_row=2, max_row=category_max_row)
    savings_rate_line.add_data(line_data, titles_from_data=True)
    savings_rate_line.set_categories(line_cats)
    savings_rate_line.height = 8
    savings_rate_line.width = 13
    ws.add_chart(savings_rate_line, "B24")

    ws["N24"] = "Data Sources"
    ws["N24"].font = Font(bold=True, color="0F172A")
    ws["N25"] = "- Benchmark_Overall"
    ws["N26"] = "- Benchmark_Category"
    ws["N27"] = "- Benchmark_Region"
    ws["N28"] = f"- Grouping keys: {category_col}, {region_col}"


def _run_pipeline() -> None:
    config = load_config()
    user = get_user("analyst1@example.com")
    authorize(user, "run_pipeline")

    input_dir = config["paths"]["input_dir"]
    output_dir = config["paths"]["output_dir"]
    os.makedirs(input_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    input_ext = config["files"]["input_extension"]
    cleaned_output = os.path.join(output_dir, config["files"]["cleaned_output"])
    report_output = os.path.join(output_dir, config["files"]["report_output"])
    benchmark_output = os.path.join(output_dir, "benchmark_summary.csv")

    qty_col = config["columns"]["quantity"]
    price_col = config["columns"]["unit_price"]
    status_col = config["columns"]["status"]
    category_col = config["columns"]["category"]
    region_col = config["columns"]["region"]
    expense_col = config.get("columns", {}).get("expense", "Expense")
    order_id_col = config.get("columns", {}).get("order_id")

    excel_files = sorted(
        [
            f for f in os.listdir(input_dir)
            if f.endswith(input_ext)
            and not f.startswith("~$")
            and os.path.isfile(os.path.join(input_dir, f))
        ]
    )

    quality_issues: list[dict] = []

    if not excel_files:
        emit_event(
            event_type="PIPELINE_SKIPPED",
            user_id=user.id,
            payload={"reason": "no_input_files"},
            level="WARNING",
        )
        print("No Excel files found.")
        raise SystemExit(2)

    emit_event(
        event_type="PIPELINE_STARTED",
        user_id=user.id,
        payload={"files_found": len(excel_files)},
    )
    notify_pipeline_started()

    cleaned_frames = []

    for file in excel_files:
        file_path = os.path.join(input_dir, file)
        try:
            df = pd.read_excel(file_path)
        except Exception as exc:
            print(f"Skipping {file}: failed to read Excel file ({exc})")
            quality_issues.append({"File": file, "Issue": f"read_error: {exc}"})
            continue

        required_cols = [qty_col, price_col, status_col, category_col, region_col]
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            print(f"Skipping {file}: missing columns {missing}")
            quality_issues.append({"File": file, "Issue": f"missing_columns: {', '.join(missing)}"})
            continue

        df = _enrich_financial_columns(
            df=df,
            qty_col=qty_col,
            price_col=price_col,
            expense_col=expense_col,
        )

        df = df[df[qty_col] >= config["filters"]["min_quantity"]]
        df = df[df[price_col] >= config["filters"]["min_unit_price"]]

        df[status_col] = df[status_col].astype(str).str.strip()
        exclude_status = set(s.strip() for s in config["filters"]["exclude_status"])
        include_status = set(s.strip() for s in config["filters"]["include_status"])
        df = df[~df[status_col].isin(exclude_status)]
        df = df[df[status_col].isin(include_status)]

        if df.empty:
            print(f"No valid rows after cleaning for {file}")
            quality_issues.append({"File": file, "Issue": "no_valid_rows_after_cleaning"})
            continue

        df = _deduplicate_cleaned_data(df, order_id_col=order_id_col)

        cleaned_frames.append(df)

    if not cleaned_frames:
        issues_output = ""
        if quality_issues:
            issues_output = _write_data_quality_issues(output_dir, quality_issues)
            print("Data quality issues written to:", issues_output)
        emit_event(
            event_type="PIPELINE_SKIPPED",
            user_id=user.id,
            payload={
                "reason": "no_valid_data",
                "issues_count": len(quality_issues),
                "issues_file": issues_output,
            },
            level="WARNING",
        )
        print("No valid data to write after processing all files.")
        raise SystemExit(2)

    cleaned_all = pd.concat(cleaned_frames, ignore_index=True)

    if os.path.exists(cleaned_output):
        try:
            existing_cleaned = pd.read_excel(cleaned_output)
            cleaned_all = pd.concat([existing_cleaned, cleaned_all], ignore_index=True)
        except Exception as exc:
            print(f"Warning: failed reading existing cleaned output. Error: {exc}")

    cleaned_all = _enrich_financial_columns(
        df=cleaned_all,
        qty_col=qty_col,
        price_col=price_col,
        expense_col=expense_col,
    )
    cleaned_all = cleaned_all[cleaned_all[qty_col] >= config["filters"]["min_quantity"]]
    cleaned_all = cleaned_all[cleaned_all[price_col] >= config["filters"]["min_unit_price"]]
    rows_before_dedup = len(cleaned_all)
    cleaned_all = _deduplicate_cleaned_data(cleaned_all, order_id_col=order_id_col)
    rows_deduped = rows_before_dedup - len(cleaned_all)

    cleaned_all.to_excel(cleaned_output, index=False)
    emit_event(
        event_type="DATA_CLEANED",
        user_id=user.id,
        payload={
            "rows_written": len(cleaned_all),
            "rows_deduped": rows_deduped,
            "output_file": cleaned_output,
        },
    )
    notify_data_cleaned(cleaned_output)

    if order_id_col and order_id_col in cleaned_all.columns:
        total_orders = cleaned_all[order_id_col].nunique()
    else:
        total_orders = len(cleaned_all)

    total_earning = float(cleaned_all["TotalRevenue"].sum())
    total_expenses = float(cleaned_all["TotalExpense"].sum())
    total_savings = float(cleaned_all["Savings"].sum())
    avg_order_value = _safe_divide(total_earning, total_orders)
    savings_rate = _safe_divide(total_savings, total_earning)

    overall_summary = pd.DataFrame([{
        "TotalOrders": total_orders,
        "TotalRevenue": total_earning,
        "TotalEarning": total_earning,
        "Expenses": total_expenses,
        "Savings": total_savings,
        "SavingsRate": savings_rate,
        "AverageOrderValue": avg_order_value,
    }])

    category_summary = (
        cleaned_all.groupby(category_col, dropna=False)
        .agg(
            TotalRevenue=("TotalRevenue", "sum"),
            TotalEarning=("TotalRevenue", "sum"),
            Expenses=("TotalExpense", "sum"),
            Savings=("Savings", "sum"),
            TotalQuantity=(qty_col, "sum"),
        )
        .reset_index()
    )
    category_summary = _add_savings_rate(category_summary, "TotalEarning", "Savings")

    if order_id_col and order_id_col in cleaned_all.columns:
        region_summary = (
            cleaned_all.groupby(region_col, dropna=False)
            .agg(
                TotalRevenue=("TotalRevenue", "sum"),
                TotalEarning=("TotalRevenue", "sum"),
                Expenses=("TotalExpense", "sum"),
                Savings=("Savings", "sum"),
                TotalOrders=(order_id_col, "nunique"),
            )
            .reset_index()
        )
    else:
        region_summary = (
            cleaned_all.groupby(region_col, dropna=False)
            .agg(
                TotalRevenue=("TotalRevenue", "sum"),
                TotalEarning=("TotalRevenue", "sum"),
                Expenses=("TotalExpense", "sum"),
                Savings=("Savings", "sum"),
                TotalOrders=("TotalRevenue", "size"),
            )
            .reset_index()
        )
    region_summary = _add_savings_rate(region_summary, "TotalEarning", "Savings")

    benchmark_overall = overall_summary[
        [
            "TotalOrders",
            "TotalEarning",
            "Expenses",
            "Savings",
            "SavingsRate",
            "AverageOrderValue",
        ]
    ]
    benchmark_category = category_summary[
        [
            category_col,
            "TotalEarning",
            "Expenses",
            "Savings",
            "SavingsRate",
            "TotalQuantity",
        ]
    ]
    benchmark_region = region_summary[
        [
            region_col,
            "TotalOrders",
            "TotalEarning",
            "Expenses",
            "Savings",
            "SavingsRate",
        ]
    ]

    benchmark_flat = pd.DataFrame(
        [
            {"Metric": "TotalEarning", "Value": total_earning},
            {"Metric": "Expenses", "Value": total_expenses},
            {"Metric": "Savings", "Value": total_savings},
            {"Metric": "SavingsRate", "Value": savings_rate},
            {"Metric": "AverageOrderValue", "Value": avg_order_value},
            {"Metric": "TotalOrders", "Value": total_orders},
        ]
    )
    benchmark_flat.to_csv(benchmark_output, index=False)
    quality_issues_output = ""
    quality_issues_df = pd.DataFrame(quality_issues, columns=["File", "Issue"])
    if not quality_issues_df.empty:
        quality_issues_output = _write_data_quality_issues(output_dir, quality_issues)

    with pd.ExcelWriter(report_output, engine="openpyxl") as writer:
        overall_summary.to_excel(writer, sheet_name="Overall_Summary", index=False)
        category_summary.to_excel(writer, sheet_name="Category_Summary", index=False)
        region_summary.to_excel(writer, sheet_name="Region_Summary", index=False)
        benchmark_overall.to_excel(writer, sheet_name="Benchmark_Overall", index=False)
        benchmark_category.to_excel(writer, sheet_name="Benchmark_Category", index=False)
        benchmark_region.to_excel(writer, sheet_name="Benchmark_Region", index=False)
        benchmark_flat.to_excel(writer, sheet_name="Benchmark_Metrics", index=False)
        if not quality_issues_df.empty:
            quality_issues_df.to_excel(writer, sheet_name="Data_Quality_Issues", index=False)
        _format_report_workbook(writer.book)
        _build_executive_dashboard(writer.book, category_col=category_col, region_col=region_col)

    emit_event(
        event_type="PIPELINE_COMPLETED",
        user_id=user.id,
        payload={
            "files_processed": len(cleaned_frames),
            "total_rows": len(cleaned_all),
            "total_earning": total_earning,
            "expenses": total_expenses,
            "savings": total_savings,
            "benchmark_file": benchmark_output,
            "quality_issues_count": len(quality_issues),
            "quality_issues_file": quality_issues_output,
        },
    )

    notify_pipeline_completed(cleaned_output, report_output)

    print("Cleaned data written to:", cleaned_output)
    print("Summary report written to:", report_output)
    print("Benchmark summary written to:", benchmark_output)
    if quality_issues_output:
        print("Data quality issues written to:", quality_issues_output)


def main() -> None:
    try:
        _run_pipeline()
    except SystemExit:
        raise
    except Exception as exc:
        emit_event(
            event_type="PIPELINE_FAILED",
            user_id="system",
            payload={"error": str(exc)},
        )
        try:
            notify_pipeline_failed(str(exc))
        except Exception as notify_exc:
            emit_event(
                event_type="PIPELINE_NOTIFY_FAILED",
                user_id="system",
                payload={"error": str(notify_exc)},
                level="WARNING",
            )
        raise


if __name__ == "__main__":
    main()
