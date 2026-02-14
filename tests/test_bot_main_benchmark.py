import sys
import shutil
import uuid
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from excel_bot.bot_main import (
    _add_savings_rate,
    _build_executive_dashboard,
    _enrich_financial_columns,
    _format_report_workbook,
    _write_data_quality_issues,
)


def test_enrich_financial_columns_uses_expense_column_when_present():
    df = pd.DataFrame(
        [
            {"Quantity": 2, "UnitPrice": 10, "Expense": 6},
            {"Quantity": 1, "UnitPrice": 5, "Expense": 1.5},
        ]
    )

    result = _enrich_financial_columns(
        df=df,
        qty_col="Quantity",
        price_col="UnitPrice",
        expense_col="Expense",
    )

    assert list(result["TotalRevenue"]) == [20, 5]
    assert list(result["TotalExpense"]) == [6, 1.5]
    assert list(result["Savings"]) == [14, 3.5]


def test_enrich_financial_columns_defaults_expense_to_zero_when_missing():
    df = pd.DataFrame(
        [
            {"Quantity": 3, "UnitPrice": 4},
            {"Quantity": 5, "UnitPrice": 2},
        ]
    )

    result = _enrich_financial_columns(
        df=df,
        qty_col="Quantity",
        price_col="UnitPrice",
        expense_col="Expense",
    )

    assert list(result["TotalRevenue"]) == [12, 10]
    assert list(result["TotalExpense"]) == [0.0, 0.0]
    assert list(result["Savings"]) == [12, 10]


def test_add_savings_rate_handles_zero_earning_without_crashing():
    df = pd.DataFrame(
        [
            {"TotalEarning": 0.0, "Savings": 0.0},
            {"TotalEarning": 100.0, "Savings": 25.0},
        ]
    )

    result = _add_savings_rate(df, earning_col="TotalEarning", savings_col="Savings")

    assert list(result["SavingsRate"]) == [0.0, 0.25]


def test_format_report_workbook_applies_metric_number_formats():
    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark_Overall"
    ws.append(["TotalOrders", "TotalEarning", "Expenses", "Savings", "SavingsRate", "AverageOrderValue"])
    ws.append([10, 1200.0, 300.0, 900.0, 0.75, 120.0])

    metrics = wb.create_sheet("Benchmark_Metrics")
    metrics.append(["Metric", "Value"])
    metrics.append(["TotalEarning", 1200.0])
    metrics.append(["SavingsRate", 0.75])
    metrics.append(["TotalOrders", 10])

    _format_report_workbook(wb)

    assert wb["Benchmark_Overall"]["B2"].number_format == "$#,##0.00"
    assert wb["Benchmark_Overall"]["E2"].number_format == "0.00%"
    assert wb["Benchmark_Metrics"]["B2"].number_format == "$#,##0.00"
    assert wb["Benchmark_Metrics"]["B3"].number_format == "0.00%"
    assert wb["Benchmark_Metrics"]["B4"].number_format == "#,##0"


def test_build_executive_dashboard_creates_visual_sheet_with_charts():
    wb = Workbook()
    overall = wb.active
    overall.title = "Benchmark_Overall"
    overall.append(["TotalOrders", "TotalEarning", "Expenses", "Savings", "SavingsRate", "AverageOrderValue"])
    overall.append([10, 1200.0, 300.0, 900.0, 0.75, 120.0])

    category = wb.create_sheet("Benchmark_Category")
    category.append(["Category", "TotalEarning", "Expenses", "Savings", "SavingsRate", "TotalQuantity"])
    category.append(["Widgets", 700.0, 200.0, 500.0, 0.7143, 40])
    category.append(["Gadgets", 500.0, 100.0, 400.0, 0.8, 25])

    region = wb.create_sheet("Benchmark_Region")
    region.append(["Region", "TotalOrders", "TotalEarning", "Expenses", "Savings", "SavingsRate"])
    region.append(["North", 6, 800.0, 220.0, 580.0, 0.725])
    region.append(["South", 4, 400.0, 80.0, 320.0, 0.8])

    _build_executive_dashboard(wb, category_col="Category", region_col="Region")

    assert "Executive_Dashboard" in wb.sheetnames
    dashboard = wb["Executive_Dashboard"]
    assert dashboard["B1"].value == "Executive Financial Dashboard"
    assert len(dashboard._charts) == 3


def test_write_data_quality_issues_creates_csv():
    base_dir = Path(__file__).resolve().parent / "tmp"
    base_dir.mkdir(exist_ok=True)
    temp_dir = base_dir / f"run_{uuid.uuid4().hex}"
    temp_dir.mkdir()

    issues = [
        {"File": "bad.xlsx", "Issue": "read_error: truncated header"},
        {"File": "missing_cols.xlsx", "Issue": "missing_columns: Quantity"},
    ]
    try:
        output_path = _write_data_quality_issues(str(temp_dir), issues)

        csv_path = Path(output_path)
        assert csv_path.exists()
        loaded = pd.read_csv(csv_path)
        assert list(loaded.columns) == ["File", "Issue"]
        assert len(loaded) == 2
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)
